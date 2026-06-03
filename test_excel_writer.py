#!/usr/bin/env python3
"""
Unit tests for Excel writer module.
"""

import pytest
import os
import tempfile
from pathlib import Path
from excel_writer import format_column_width, create_backup, auto_adjust_columns
from openpyxl import Workbook


class TestFormatColumnWidth:
    """Test column width formatting."""
    
    def test_format_simple_length(self):
        """Test formatting simple column width."""
        width = format_column_width(10)
        assert width == 12  # 10 + 2 padding
    
    def test_format_long_length(self):
        """Test formatting very long column."""
        width = format_column_width(100)
        assert width == 80  # Capped at max
    
    def test_format_zero_length(self):
        """Test formatting empty column."""
        width = format_column_width(0)
        assert width == 2


class TestCreateBackup:
    """Test backup file creation."""
    
    def test_backup_existing_file(self):
        """Test creating backup of existing file."""
        with tempfile.NamedTemporaryFile(delete=False) as f:
            temp_path = f.name
            f.write(b"test content")
        
        try:
            # Mock config
            from config import Config
            original_backup = Config.BACKUP_FILES
            Config.BACKUP_FILES = True
            
            backup_path = create_backup(temp_path)
            
            # Restore config
            Config.BACKUP_FILES = original_backup
            
            assert backup_path is not None
            assert os.path.exists(backup_path)
            assert os.path.exists(temp_path)
            
            # Clean up
            os.remove(backup_path)
        finally:
            os.remove(temp_path)
    
    def test_backup_nonexistent_file(self):
        """Test backup of non-existent file."""
        backup_path = create_backup("/nonexistent/file.xlsx")
        assert backup_path is None
    
    def test_backup_disabled(self):
        """Test backup when disabled."""
        with tempfile.NamedTemporaryFile(delete=False) as f:
            temp_path = f.name
        
        try:
            # Mock config
            from config import Config
            original_backup = Config.BACKUP_FILES
            Config.BACKUP_FILES = False
            
            backup_path = create_backup(temp_path)
            
            # Restore config
            Config.BACKUP_FILES = original_backup
            
            assert backup_path is None
        finally:
            os.remove(temp_path)


class TestAutoAdjustColumns:
    """Test auto column adjustment."""
    
    def test_adjust_columns_simple(self):
        """Test adjusting columns with simple data."""
        wb = Workbook()
        ws = wb.active
        
        # Add some data
        ws.append(["Short", "Medium Text", "Very Long Text Column"])
        ws.append(["A", "BB", "CCC"])
        
        # Adjust columns
        auto_adjust_columns(ws)
        
        # Check that columns have been resized
        assert ws.column_dimensions['A'].width > 0
        assert ws.column_dimensions['B'].width > 0
        assert ws.column_dimensions['C'].width > 0
    
    def test_adjust_columns_empty(self):
        """Test adjusting empty worksheet."""
        wb = Workbook()
        ws = wb.active
        
        # Should not raise error
        auto_adjust_columns(ws)
    
    def test_adjust_columns_respects_max_width(self):
        """Test that column width respects maximum."""
        wb = Workbook()
        ws = wb.active
        
        # Add very long data
        long_text = "A" * 200
        ws.append([long_text])
        
        # Adjust columns
        auto_adjust_columns(ws, max_width=50)
        
        # Check that max width is respected
        assert ws.column_dimensions['A'].width <= 50


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
