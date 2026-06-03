#!/usr/bin/env python3
"""
Excel file writing utilities for bank statement data.
Handles Excel workbook creation, formatting, and safe file operations.
"""

import os
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from config import Config, get_logger

logger = get_logger('mutasi.excel')


def format_column_width(max_length: int) -> float:
    """Calculate Excel column width from content length.
    
    Args:
        max_length: Maximum content length in column
        
    Returns:
        Width value for Excel column
    """
    return min(max_length + Config.MIN_COLUMN_WIDTH, Config.MAX_COLUMN_WIDTH)


def create_backup(output_path: str) -> Optional[str]:
    """Create backup of existing file.
    
    Args:
        output_path: Path to file to backup
        
    Returns:
        Path to backup file or None if not created
    """
    if not os.path.exists(output_path):
        return None
    
    if not Config.BACKUP_FILES:
        logger.debug(f"Backup disabled, skipping: {output_path}")
        return None
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{output_path}.backup.{timestamp}"
        shutil.copy2(output_path, backup_path)
        logger.info(f"Created backup: {backup_path}")
        return backup_path
    except Exception as e:
        logger.warning(f"Failed to create backup: {e}")
        return None


def auto_adjust_columns(worksheet: Worksheet, max_width: int = Config.MAX_COLUMN_WIDTH) -> None:
    """Auto-adjust column widths based on content.
    
    Args:
        worksheet: Worksheet to adjust
        max_width: Maximum width for columns
    """
    try:
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            
            for cell in col:
                try:
                    cell_value = str(cell.value or '')
                    if len(cell_value) > max_len:
                        max_len = len(cell_value)
                except (AttributeError, TypeError):
                    continue
            
            # Set column width
            worksheet.column_dimensions[col_letter].width = format_column_width(max_len)
            
    except Exception as e:
        logger.error(f"Error adjusting column widths: {e}")


def save_excel_file(workbook: Workbook, output_path: str) -> None:
    """Save Excel workbook with atomic operation.
    
    Args:
        workbook: Workbook to save
        output_path: Path where to save file
        
    Raises:
        IOError: If save operation fails
    """
    try:
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Create backup if file exists
        create_backup(output_path)
        
        # Save to temporary file first (atomic operation)
        temp_path = output_path + '.tmp'
        workbook.save(temp_path)
        
        # Verify temp file was created and has content
        if not os.path.exists(temp_path) or os.path.getsize(temp_path) == 0:
            raise IOError("Saved file is empty or missing")
        
        # Atomic move
        os.replace(temp_path, output_path)
        logger.info(f"Successfully saved Excel file: {output_path}")
        
    except Exception as e:
        # Clean up temp file if it exists
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
        logger.error(f"Failed to save Excel file: {e}")
        raise


def write_transactions_to_excel(
    transactions: List[Dict],
    output_path: str,
    sheet_name: str = "Mutasi Rekening"
) -> None:
    """Write transactions to Excel file.
    
    Args:
        transactions: List of transaction dictionaries
        output_path: Path to output Excel file
        sheet_name: Name for the worksheet
        
    Raises:
        IOError: If file operation fails
    """
    try:
        wb = Workbook()
        ws = wb.active
        
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        
        ws.title = sheet_name
        
        # Write headers
        headers = ['Tanggal', 'Bulan', 'Keterangan', 'DB', 'CR', 'Saldo']
        ws.append(headers)
        
        # Write transactions
        row_count = 0
        for tx in transactions:
            try:
                ws.append([
                    tx.get('tanggal', ''),
                    tx.get('bulan', ''),
                    tx.get('keterangan', ''),
                    tx.get('db', ''),
                    tx.get('cr', ''),
                    tx.get('saldo', '')
                ])
                row_count += 1
            except Exception as e:
                logger.warning(f"Skipped invalid transaction: {e}")
                continue
        
        # Auto-adjust column widths
        auto_adjust_columns(ws)
        
        # Save file
        save_excel_file(wb, output_path)
        logger.info(f"Wrote {row_count} transactions to {output_path}")
        
    except Exception as e:
        logger.error(f"Error writing to Excel: {e}")
        raise


def write_multiple_sheets_to_excel(
    transactions_by_sheet: Dict[str, List[Dict]],
    output_path: str
) -> None:
    """Write multiple sheets to single Excel file.
    
    Args:
        transactions_by_sheet: Dictionary of sheet_name -> transactions
        output_path: Path to output Excel file
        
    Raises:
        IOError: If file operation fails
    """
    try:
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
        
        total_rows = 0
        
        for sheet_name, transactions in transactions_by_sheet.items():
            try:
                ws = wb.create_sheet(title=sheet_name)
                
                # Write headers
                headers = ['Tanggal', 'Bulan', 'Keterangan', 'DB', 'CR', 'Saldo']
                ws.append(headers)
                
                # Write transactions
                row_count = 0
                for tx in transactions:
                    try:
                        ws.append([
                            tx.get('tanggal', ''),
                            tx.get('bulan', ''),
                            tx.get('keterangan', ''),
                            tx.get('db', ''),
                            tx.get('cr', ''),
                            tx.get('saldo', '')
                        ])
                        row_count += 1
                    except Exception as e:
                        logger.warning(f"Skipped invalid transaction in {sheet_name}: {e}")
                        continue
                
                # Auto-adjust column widths
                auto_adjust_columns(ws)
                total_rows += row_count
                logger.info(f"Sheet '{sheet_name}': {row_count} transactions")
                
            except Exception as e:
                logger.error(f"Error creating sheet '{sheet_name}': {e}")
                continue
        
        # Save file
        save_excel_file(wb, output_path)
        logger.info(f"Wrote {total_rows} total transactions across {len(transactions_by_sheet)} sheets")
        
    except Exception as e:
        logger.error(f"Error writing multiple sheets: {e}")
        raise
